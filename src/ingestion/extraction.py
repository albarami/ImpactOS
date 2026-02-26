"""Document extraction service — MVP-2 Sections 8.2, 8.3, 8.5.

Extracts tables from CSV and Excel files, builds a DocumentGraph with
bounding boxes, and generates EvidenceSnippets per data row.

This is a deterministic service — no LLM calls. PDF extraction uses
a layout-aware approach; CSV/Excel bypass OCR and parse directly
(Section 8.2: "Structured inputs bypass OCR").
"""

import csv
import io
from uuid import UUID

import openpyxl

from src.models.common import utc_now
from src.models.document import (
    DocumentGraph,
    ExtractionMetadata,
    ExtractedTable,
    PageBlock,
    TableCell,
    TextBlock,
)
from src.models.governance import BoundingBox, EvidenceSnippet, TableCellRef


class ExtractionService:
    """Deterministic document extraction engine.

    Supports CSV and Excel. PDF support would use an external layout-aware
    service (Section 8.2) and is not yet implemented.
    """

    # ------------------------------------------------------------------
    # CSV extraction
    # ------------------------------------------------------------------

    def extract_csv(
        self,
        *,
        doc_id: UUID,
        content: bytes,
        doc_checksum: str,
    ) -> DocumentGraph:
        """Extract a single table from CSV content.

        CSV is treated as a single-page, single-table document.
        Bounding boxes are synthesized from row/col positions.
        Confidence is 1.0 (exact data).
        """
        text = content.decode("utf-8")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)

        num_rows = len(rows)
        num_cols = max((len(r) for r in rows), default=0)

        cells = self._rows_to_cells(rows, num_rows, num_cols)

        table = ExtractedTable(
            table_id="csv_table_0",
            page_number=0,
            bbox=BoundingBox(x0=0.0, y0=0.0, x1=1.0, y1=1.0),
            cells=cells,
        )

        page = PageBlock(
            page_number=0,
            blocks=[],
            tables=[table],
        )

        return DocumentGraph(
            document_id=doc_id,
            pages=[page],
            extraction_metadata=ExtractionMetadata(
                engine="csv-parser",
                engine_version="1.0.0",
                completed_at=utc_now(),
            ),
        )

    # ------------------------------------------------------------------
    # Excel extraction
    # ------------------------------------------------------------------

    def extract_excel(
        self,
        *,
        doc_id: UUID,
        content: bytes,
        doc_checksum: str,
    ) -> DocumentGraph:
        """Extract tables from Excel workbook (one table per sheet).

        Each sheet becomes a page. Bounding boxes are synthesized.
        Confidence is 1.0 (exact data).
        """
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        pages: list[PageBlock] = []

        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            rows: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(v) if v is not None else "" for v in row])

            num_rows = len(rows)
            num_cols = max((len(r) for r in rows), default=0)

            cells = self._rows_to_cells(rows, num_rows, num_cols)

            table = ExtractedTable(
                table_id=f"xlsx_{sheet_name}_{sheet_idx}",
                page_number=sheet_idx,
                bbox=BoundingBox(x0=0.0, y0=0.0, x1=1.0, y1=1.0),
                cells=cells,
            )

            page = PageBlock(
                page_number=sheet_idx,
                blocks=[],
                tables=[table],
            )
            pages.append(page)

        wb.close()

        return DocumentGraph(
            document_id=doc_id,
            pages=pages,
            extraction_metadata=ExtractionMetadata(
                engine="openpyxl",
                engine_version="1.0.0",
                completed_at=utc_now(),
            ),
        )

    # ------------------------------------------------------------------
    # Evidence snippet generation (Section 8.5)
    # ------------------------------------------------------------------

    def generate_evidence_snippets(
        self,
        *,
        document_graph: DocumentGraph,
        source_id: UUID,
        doc_checksum: str,
    ) -> list[EvidenceSnippet]:
        """Generate one EvidenceSnippet per data row (skipping header row 0).

        Each snippet stores the page number, bounding box covering the row,
        extracted text (concatenation of cell values), and a TableCellRef
        pointing to the first cell in the row.
        """
        snippets: list[EvidenceSnippet] = []

        for page in document_graph.pages:
            for table in page.tables:
                # Group cells by row
                rows_map: dict[int, list[TableCell]] = {}
                for cell in table.cells:
                    rows_map.setdefault(cell.row, []).append(cell)

                # Skip row 0 (header), generate snippet per data row
                data_row_indices = sorted(r for r in rows_map if r > 0)
                for row_idx in data_row_indices:
                    row_cells = sorted(rows_map[row_idx], key=lambda c: c.col)
                    if not row_cells:
                        continue

                    # Merge bounding boxes across all cells in the row
                    x0 = min(c.bbox.x0 for c in row_cells)
                    y0 = min(c.bbox.y0 for c in row_cells)
                    x1 = max(c.bbox.x1 for c in row_cells)
                    y1 = max(c.bbox.y1 for c in row_cells)

                    text = " | ".join(c.text for c in row_cells)

                    snippet = EvidenceSnippet(
                        source_id=source_id,
                        page=page.page_number,
                        bbox=BoundingBox(x0=x0, y0=y0, x1=x1, y1=y1),
                        extracted_text=text,
                        table_cell_ref=TableCellRef(
                            table_id=table.table_id,
                            row=row_idx,
                            col=row_cells[0].col,
                        ),
                        checksum=doc_checksum,
                    )
                    snippets.append(snippet)

        return snippets

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _rows_to_cells(
        rows: list[list[str]],
        num_rows: int,
        num_cols: int,
    ) -> list[TableCell]:
        """Convert row/column data to TableCell objects with synthesized bboxes."""
        cells: list[TableCell] = []
        if num_rows == 0 or num_cols == 0:
            return cells

        row_height = 1.0 / num_rows
        col_width = 1.0 / num_cols

        for r_idx, row in enumerate(rows):
            for c_idx, value in enumerate(row):
                y0 = r_idx * row_height
                y1 = y0 + row_height
                x0 = c_idx * col_width
                x1 = x0 + col_width

                # Clamp to [0, 1] for floating-point safety
                x1 = min(x1, 1.0)
                y1 = min(y1, 1.0)

                cells.append(
                    TableCell(
                        row=r_idx,
                        col=c_idx,
                        text=value,
                        bbox=BoundingBox(x0=x0, y0=y0, x1=x1, y1=y1),
                        confidence=1.0,
                    )
                )

        return cells
