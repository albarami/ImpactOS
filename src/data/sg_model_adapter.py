"""SG Model Adapter -- extracts IO model artifacts from SG workbooks.

Separate from sg_template_parser.py (which handles INTERVENTIONS only).
This module parses IO_MODEL, FINAL_DEMAND, IMPORTS, and VALUE_ADDED sheets
and returns IOModelData.

CRITICAL: No LLM calls. This is a deterministic extraction module.
"""

from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import numpy as np

from src.data.io_loader import IOModelData


class SGImportError(ValueError):
    """Error during SG workbook import with a stable reason code."""

    def __init__(self, reason_code: str, message: str) -> None:
        super().__init__(f"{reason_code}: {message}")
        self.reason_code = reason_code
        self.message = message


@dataclass(frozen=True)
class SGSheetLayout:
    """Detected layout of an SG IO model workbook."""

    z_sheet: str
    x_row: int
    sector_codes_row: int
    sector_count: int
    sector_col: int
    code_col: int
    name_col: int
    data_start_row: int
    base_year: int | None
    final_demand_sheet: str | None
    imports_sheet: str | None
    value_added_sheet: str | None


def _normalize_sector_code(raw: str | int | float | None) -> str:
    """Normalize sector codes: 10.0 -> 10, 6 -> 06, 06 -> 06."""
    if raw is None:
        return ""
    if isinstance(raw, float):
        raw = int(raw)
    s = str(raw).strip()
    if s.isdigit():
        return f"{int(s):02d}"
    m = re.search(r"(\d+)", s)
    if m:
        return f"{int(m.group(1)):02d}"
    return s


def _read_all_sheets_xlsx(path: str) -> dict[str, list[list[Any]]]:
    """Read ALL sheets from .xlsx using openpyxl."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheets: dict[str, list[list[Any]]] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        sheets[name] = rows
    wb.close()
    return sheets


def _read_all_sheets_xlsb(path: str) -> dict[str, list[list[Any]]]:
    """Read ALL sheets from .xlsb using pyxlsb."""
    import pyxlsb

    wb = pyxlsb.open_workbook(path)
    sheets: dict[str, list[list[Any]]] = {}
    for name in wb.sheets:
        rows: list[list[Any]] = []
        with wb.get_sheet(name) as sheet:
            for row in sheet.rows():
                rows.append([cell.v for cell in row])
        sheets[name] = rows
    return sheets


def _compute_file_sha256(path: Path) -> str:
    """Compute SHA-256 hash of file contents, prefixed with sha256:."""
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return f"sha256:{h.hexdigest()}"


def _find_io_model_sheet(sheet_names: list[str]) -> str | None:
    """Find the IO model sheet by name convention."""
    for name in sheet_names:
        if name.upper() == "IO_MODEL":
            return name
    for name in sheet_names:
        if name.upper() == "MODEL" and "INTERVENTION" not in name.upper():
            return name
    return None


def _find_extended_sheet(sheet_names: list[str], keyword: str) -> str | None:
    """Find an extended artifact sheet by keyword match."""
    for name in sheet_names:
        if keyword.upper() in name.upper():
            return name
    return None


def detect_sg_layout(path: Path) -> SGSheetLayout:
    """Detect the layout of an SG IO model workbook.

    Raises:
        SGImportError: SG_FILE_UNREADABLE if file cannot be opened.
        SGImportError: SG_LAYOUT_DETECTION_FAILED if structure not found.
    """
    path = Path(path)

    if not path.exists():
        raise SGImportError("SG_FILE_UNREADABLE", f"File not found: {path}")

    ext = path.suffix.lower()
    try:
        if ext == ".xlsb":
            all_sheets = _read_all_sheets_xlsb(str(path))
        elif ext in (".xlsx", ".xls"):
            all_sheets = _read_all_sheets_xlsx(str(path))
        else:
            raise SGImportError(
                "SG_FILE_UNREADABLE",
                f"Unsupported file extension: {ext}",
            )
    except SGImportError:
        raise
    except Exception as exc:
        raise SGImportError(
            "SG_FILE_UNREADABLE",
            f"Cannot read workbook: {exc}",
        ) from exc

    sheet_names = list(all_sheets.keys())
    io_sheet_name = _find_io_model_sheet(sheet_names)

    # Fallback: scan all non-INTERVENTION sheets for CODE header
    if io_sheet_name is None:
        for name in sheet_names:
            if "INTERVENTION" in name.upper():
                continue
            rows = all_sheets[name]
            for row in rows:
                for val in row:
                    if isinstance(val, str) and val.strip().upper() == "CODE":
                        io_sheet_name = name
                        break
                if io_sheet_name:
                    break
            if io_sheet_name:
                break

    if io_sheet_name is None:
        raise SGImportError(
            "SG_LAYOUT_DETECTION_FAILED",
            "No IO model sheet found.",
        )

    rows = all_sheets[io_sheet_name]
    code_row_idx: int | None = None
    code_col: int = 0
    name_col: int = 1

    for ri, row in enumerate(rows):
        for ci, val in enumerate(row):
            if isinstance(val, str) and val.strip().upper() == "CODE":
                code_row_idx = ri
                code_col = ci
                name_col = ci + 1
                break
        if code_row_idx is not None:
            break

    if code_row_idx is None:
        raise SGImportError(
            "SG_LAYOUT_DETECTION_FAILED",
            "No CODE header found in sheet.",
        )

    header_row = rows[code_row_idx]
    sector_col = name_col + 1
    sector_codes_in_header: list[str] = []
    for ci in range(sector_col, len(header_row)):
        val = header_row[ci]
        if val is None:
            break
        code = _normalize_sector_code(val)
        if code and code[0].isdigit():
            sector_codes_in_header.append(code)
        else:
            break

    sector_count = len(sector_codes_in_header)
    if sector_count == 0:
        raise SGImportError(
            "SG_LAYOUT_DETECTION_FAILED",
            "No sector codes found in header row.",
        )

    x_row: int | None = None
    data_start_row = code_row_idx + 1

    for ri in range(data_start_row, len(rows)):
        row = rows[ri]
        if not row or len(row) <= code_col:
            continue
        first_cell = row[code_col]
        if isinstance(first_cell, str) and "TOTAL_OUTPUT" in first_cell.upper():
            x_row = ri
            break

    if x_row is None:
        raise SGImportError(
            "SG_LAYOUT_DETECTION_FAILED",
            "TOTAL_OUTPUT row not found.",
        )

    base_year: int | None = None
    for row in rows:
        for ci, val in enumerate(row):
            if isinstance(val, str) and "BASE_YEAR" in val.upper():
                if ci + 1 < len(row) and row[ci + 1] is not None:
                    try:
                        base_year = int(row[ci + 1])
                    except (ValueError, TypeError):
                        pass
                break

    fd_sheet = _find_extended_sheet(sheet_names, "FINAL_DEMAND")
    imp_sheet = _find_extended_sheet(sheet_names, "IMPORT")
    va_sheet = _find_extended_sheet(sheet_names, "VALUE_ADDED")

    return SGSheetLayout(
        z_sheet=io_sheet_name,
        x_row=x_row,
        sector_codes_row=code_row_idx,
        sector_count=sector_count,
        sector_col=sector_col,
        code_col=code_col,
        name_col=name_col,
        data_start_row=data_start_row,
        base_year=base_year,
        final_demand_sheet=fd_sheet,
        imports_sheet=imp_sheet,
        value_added_sheet=va_sheet,
    )


def _extract_sector_info(
    rows: list[list[Any]],
    layout: SGSheetLayout,
) -> tuple[list[str], dict[str, str]]:
    """Extract sector codes and names from the IO model sheet."""
    codes: list[str] = []
    names: dict[str, str] = {}

    for ri in range(layout.data_start_row, layout.data_start_row + layout.sector_count):
        if ri >= len(rows):
            break
        row = rows[ri]
        raw_code = row[layout.code_col] if layout.code_col < len(row) else None
        code = _normalize_sector_code(raw_code)
        if not code:
            continue
        codes.append(code)

        raw_name = row[layout.name_col] if layout.name_col < len(row) else None
        if raw_name is not None:
            names[code] = str(raw_name).strip()

    return codes, names


def _safe_float(val: object) -> float:
    """Convert a cell value to float, defaulting to NaN on failure."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return float("nan")


def _extract_z_matrix(
    rows: list[list[Any]],
    layout: SGSheetLayout,
) -> np.ndarray:
    """Extract the Z (intermediate flow) matrix."""
    n = layout.sector_count
    z = np.zeros((n, n), dtype=np.float64)

    for i in range(n):
        ri = layout.data_start_row + i
        if ri >= len(rows):
            raise SGImportError(
                "SG_PARSE_MATRIX_FAILED",
                f"Row {ri} missing for sector {i}.",
            )
        row = rows[ri]
        for j in range(n):
            ci = layout.sector_col + j
            if ci >= len(row):
                raise SGImportError(
                    "SG_PARSE_MATRIX_FAILED",
                    f"Column {ci} missing at row {ri}.",
                )
            val = _safe_float(row[ci])
            if np.isnan(val):
                raise SGImportError(
                    "SG_PARSE_MATRIX_FAILED",
                    f"Non-numeric value at row {ri}, col {ci}: {row[ci]!r}",
                )
            z[i, j] = val

    return z


def _extract_x_vector(
    rows: list[list[Any]],
    layout: SGSheetLayout,
) -> np.ndarray:
    """Extract the total output (x) vector from the TOTAL_OUTPUT row."""
    n = layout.sector_count
    x = np.zeros(n, dtype=np.float64)

    if layout.x_row >= len(rows):
        raise SGImportError(
            "SG_PARSE_MATRIX_FAILED",
            f"TOTAL_OUTPUT row {layout.x_row} out of range.",
        )

    row = rows[layout.x_row]
    for j in range(n):
        ci = layout.sector_col + j
        if ci >= len(row):
            raise SGImportError(
                "SG_PARSE_MATRIX_FAILED",
                f"TOTAL_OUTPUT column {ci} missing.",
            )
        val = _safe_float(row[ci])
        if np.isnan(val):
            raise SGImportError(
                "SG_PARSE_MATRIX_FAILED",
                f"Non-numeric TOTAL_OUTPUT at col {ci}: {row[ci]!r}",
            )
        x[j] = val

    return x


def _extract_extended_vector(
    rows: list[list[Any]],
    n_sectors: int,
    col_idx: int,
) -> np.ndarray:
    """Extract a single numeric column vector from an extended sheet."""
    vec = np.zeros(n_sectors, dtype=np.float64)
    for i in range(n_sectors):
        ri = i + 1  # skip header
        if ri >= len(rows):
            raise SGImportError(
                "SG_PARSE_ARTIFACT_FAILED",
                f"Extended sheet row {ri} missing for sector {i}.",
            )
        row = rows[ri]
        if col_idx >= len(row):
            raise SGImportError(
                "SG_PARSE_ARTIFACT_FAILED",
                f"Extended sheet col {col_idx} missing at row {ri}.",
            )
        val = _safe_float(row[col_idx])
        if np.isnan(val):
            raise SGImportError(
                "SG_PARSE_ARTIFACT_FAILED",
                f"Non-numeric value at row {ri}, col {col_idx}: {row[col_idx]!r}",
            )
        vec[i] = val
    return vec


def _extract_extended_matrix(
    rows: list[list[Any]],
    n_sectors: int,
    col_start: int,
    n_cols: int,
) -> np.ndarray:
    """Extract a (n_sectors x n_cols) matrix from an extended sheet."""
    mat = np.zeros((n_sectors, n_cols), dtype=np.float64)
    for i in range(n_sectors):
        ri = i + 1
        if ri >= len(rows):
            raise SGImportError(
                "SG_PARSE_ARTIFACT_FAILED",
                f"Extended sheet row {ri} missing.",
            )
        row = rows[ri]
        for j in range(n_cols):
            ci = col_start + j
            if ci >= len(row):
                raise SGImportError(
                    "SG_PARSE_ARTIFACT_FAILED",
                    f"Extended sheet col {ci} missing at row {ri}.",
                )
            val = _safe_float(row[ci])
            if np.isnan(val):
                raise SGImportError(
                    "SG_PARSE_ARTIFACT_FAILED",
                    f"Non-numeric value at row {ri}, col {ci}: {row[ci]!r}",
                )
            mat[i, j] = val
    return mat


def extract_io_model(
    path: Path,
    *,
    layout: SGSheetLayout | None = None,
) -> IOModelData:
    """Extract IO model data from an SG workbook.

    Args:
        path: Path to .xlsx or .xlsb workbook.
        layout: Pre-detected layout (if None, calls detect_sg_layout).

    Returns:
        IOModelData with Z, x, sector_codes, sector_names, and extended
        artifacts if present.

    Raises:
        SGImportError: With specific reason codes.
    """
    path = Path(path)

    if not path.exists():
        raise SGImportError("SG_FILE_UNREADABLE", f"File not found: {path}")

    if layout is None:
        layout = detect_sg_layout(path)

    ext = path.suffix.lower()
    try:
        if ext == ".xlsb":
            all_sheets = _read_all_sheets_xlsb(str(path))
        elif ext in (".xlsx", ".xls"):
            all_sheets = _read_all_sheets_xlsx(str(path))
        else:
            raise SGImportError(
                "SG_FILE_UNREADABLE",
                f"Unsupported file extension: {ext}",
            )
    except SGImportError:
        raise
    except Exception as exc:
        raise SGImportError(
            "SG_FILE_UNREADABLE",
            f"Cannot read workbook: {exc}",
        ) from exc

    io_rows = all_sheets.get(layout.z_sheet)
    if io_rows is None:
        raise SGImportError(
            "SG_PARSE_MATRIX_FAILED",
            f"Sheet not found in workbook: {layout.z_sheet}",
        )

    # Extract sectors
    try:
        sector_codes, sector_names = _extract_sector_info(io_rows, layout)
    except SGImportError:
        raise
    except Exception as exc:
        raise SGImportError(
            "SG_PARSE_SECTORS_FAILED",
            f"Failed to extract sector info: {exc}",
        ) from exc

    if len(sector_codes) != layout.sector_count:
        raise SGImportError(
            "SG_PARSE_SECTORS_FAILED",
            f"Expected {layout.sector_count} sectors, found {len(sector_codes)}.",
        )

    # Extract Z matrix
    try:
        z_matrix = _extract_z_matrix(io_rows, layout)
    except SGImportError:
        raise
    except Exception as exc:
        raise SGImportError(
            "SG_PARSE_MATRIX_FAILED",
            f"Failed to extract Z matrix: {exc}",
        ) from exc

    # Extract x vector
    try:
        x_vector = _extract_x_vector(io_rows, layout)
    except SGImportError:
        raise
    except Exception as exc:
        raise SGImportError(
            "SG_PARSE_MATRIX_FAILED",
            f"Failed to extract x vector: {exc}",
        ) from exc

    # Compute file hash
    file_hash = _compute_file_sha256(path)

    # Extract extended artifacts
    n = layout.sector_count
    final_demand_f: np.ndarray | None = None
    imports_vec: np.ndarray | None = None
    compensation: np.ndarray | None = None
    gos: np.ndarray | None = None
    taxes: np.ndarray | None = None

    try:
        if layout.final_demand_sheet and layout.final_demand_sheet in all_sheets:
            fd_rows = all_sheets[layout.final_demand_sheet]
            if fd_rows:
                header = fd_rows[0]
                n_demand_cols = len([v for v in header[2:] if v is not None])
                if n_demand_cols > 0:
                    final_demand_f = _extract_extended_matrix(
                        fd_rows, n, col_start=2, n_cols=n_demand_cols,
                    )

        if layout.imports_sheet and layout.imports_sheet in all_sheets:
            imp_rows = all_sheets[layout.imports_sheet]
            if imp_rows:
                imports_vec = _extract_extended_vector(imp_rows, n, col_idx=2)

        if layout.value_added_sheet and layout.value_added_sheet in all_sheets:
            va_rows = all_sheets[layout.value_added_sheet]
            if va_rows:
                compensation = _extract_extended_vector(va_rows, n, col_idx=2)
                gos = _extract_extended_vector(va_rows, n, col_idx=3)
                taxes = _extract_extended_vector(va_rows, n, col_idx=4)

    except SGImportError:
        raise
    except Exception as exc:
        raise SGImportError(
            "SG_PARSE_ARTIFACT_FAILED",
            f"Failed to extract extended artifacts: {exc}",
        ) from exc

    provenance = {
        "workbook_sha256": file_hash,
        "source_filename": path.name,
        "import_mode": "sg_workbook",
        "imported_at": datetime.now(tz=UTC).isoformat(),
    }

    return IOModelData(
        Z=z_matrix,
        x=x_vector,
        sector_codes=sector_codes,
        sector_names=sector_names,
        base_year=layout.base_year or 0,
        source=f"sg_workbook:{path.name}",
        metadata={"sg_provenance": provenance},
        final_demand_F=final_demand_f,
        imports_vector=imports_vec,
        compensation_of_employees=compensation,
        gross_operating_surplus=gos,
        taxes_less_subsidies=taxes,
    )
