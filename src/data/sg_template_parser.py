"""Parse Strategic Gears production Excel templates (.xlsb or .xlsx).

Extracts intervention data from SG's production workbook into structured
dataclasses suitable for engine consumption.

CRITICAL: Column positions are NOT hardcoded. The parser detects structure
dynamically by scanning for 'CODE', 'SECTOR', and year column patterns.

Supports:
    .xlsb — via pyxlsb (SG production format)
    .xlsx — via openpyxl (export/converted format)
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import numpy as np

from src.data.concordance import ConcordanceService

# =====================================================================
# Data models
# =====================================================================


@dataclass(frozen=True)
class SGShockRow:
    """Single sector-row from an SG intervention block."""

    intervention_label: str
    raw_change_type: str
    normalized_change_type: str
    raw_measure_label: str
    normalized_measure: str
    sector_code: str
    sector_name: str
    is_selected: bool
    annual_values: dict[int, float]


@dataclass(frozen=True)
class SGIntervention:
    """Grouped intervention from SG template."""

    label: str
    change_type: str
    measure: str
    rows: list[SGShockRow]
    active_rows: list[SGShockRow]


@dataclass(frozen=True)
class SGScenarioImport:
    """Complete parsed scenario from SG template."""

    scenario_name: str
    interventions: list[SGIntervention]
    years: list[int]
    all_sector_codes: list[str]
    metadata: dict[str, object]


@dataclass(frozen=True)
class TranslationReport:
    """Classifies which interventions can be executed by current engine."""

    supported: list[SGIntervention]
    unsupported: list[SGIntervention]
    warnings: list[str]


# =====================================================================
# Normalization functions
# =====================================================================


def normalize_sector_code(raw: str | int | float | None) -> str:
    """Normalize sector codes from various Excel formats.

    10.0 -> "10"
    "06" -> "06"
    " 06 " -> "06"
    6 -> "06"
    6.0 -> "06"
    """
    if raw is None:
        return ""
    if isinstance(raw, float):
        raw = int(raw)
    s = str(raw).strip()
    # Remove leading/trailing whitespace already done
    # If numeric, zero-pad to 2 digits
    if s.isdigit():
        return f"{int(s):02d}"
    # Handle " - 06" format from B columns
    m = re.search(r"(\d+)", s)
    if m:
        return f"{int(m.group(1)):02d}"
    return s


def normalize_change_type(raw: str) -> str:
    """Normalize change type labels.

    "Absolute_Value" -> "absolute_value"
    "Absolute Value" -> "absolute_value"
    "  Percentage " -> "percentage"
    "Sector Loss" -> "sector_loss"
    """
    s = str(raw).strip().lower()
    s = s.replace(" ", "_")
    # Collapse multiple underscores
    s = re.sub(r"_+", "_", s)
    return s


def _normalize_measure(raw: str) -> str:
    """Normalize intervention measure labels.

    "" or None -> "final_demand"
    "Compensation of employees" -> "compensation_of_employees"
    """
    if not raw or not str(raw).strip():
        return "final_demand"
    s = str(raw).strip().lower()
    s = s.replace(" ", "_")
    s = re.sub(r"_+", "_", s)
    return s


def parse_parenthetical_negative(raw: str | int | float) -> float:
    """Parse Excel parenthetical negatives.

    (0.05) -> -0.05
    0.05 -> 0.05
    "-0.05" -> -0.05
    """
    if isinstance(raw, int | float):
        return float(raw)
    s = str(raw).strip()
    # Match (number) pattern
    m = re.match(r"^\(([0-9.]+)\)$", s)
    if m:
        return -float(m.group(1))
    return float(s)


# =====================================================================
# Grid detection
# =====================================================================


def _find_header_row(
    rows: list[list[Any]],
) -> tuple[int, list[dict[str, Any]]]:
    """Find the header row and detect intervention blocks.

    Scans for rows containing 'CODE' and year values (2020-2035).
    Returns (header_row_index, list_of_intervention_block_defs).

    Each block def: {
        'code_col': int,
        'sector_col': int,
        'year_cols': {year: col_idx, ...},
        'change_type_col': int | None,
    }
    """
    for row_idx, row_vals in enumerate(rows):
        # Find CODE column
        code_col = None
        for ci, val in enumerate(row_vals):
            if isinstance(val, str) and val.strip().upper() == "CODE":
                code_col = ci
                break

        if code_col is None:
            continue

        # Found CODE - now find year columns (look for integers 2020-2035)
        # IMPORTANT: Use list of (col, year) pairs — NOT a dict — because
        # multiple intervention blocks share the same years (2025-2030).
        # A dict would lose duplicates.
        year_col_pairs: list[tuple[int, int]] = []  # (col_idx, year)
        for ci, val in enumerate(row_vals):
            if isinstance(val, int | float):
                yr = int(val)
                if 2020 <= yr <= 2035:
                    year_col_pairs.append((ci, yr))

        if not year_col_pairs:
            continue

        # Found header row. SECTOR is typically CODE+1
        sector_col = code_col + 1

        # Sort by column index for gap detection
        year_col_pairs.sort(key=lambda p: p[0])

        # Split into contiguous groups (separated by column gaps > 2)
        blocks: list[dict[str, Any]] = []
        current_pairs: list[tuple[int, int]] = []

        for ci, yr in year_col_pairs:
            if current_pairs:
                last_col = current_pairs[-1][0]
                # If there's a gap > 2 columns, start new block
                if ci - last_col > 2:
                    blocks.append({
                        "code_col": code_col,
                        "sector_col": sector_col,
                        "year_cols": {yr_: col_ for col_, yr_ in current_pairs},
                    })
                    current_pairs = []
            current_pairs.append((ci, yr))

        if current_pairs:
            blocks.append({
                "code_col": code_col,
                "sector_col": sector_col,
                "year_cols": {yr_: col_ for col_, yr_ in current_pairs},
            })

        return row_idx, blocks

    raise ValueError(
        "Could not find header row with CODE and year columns. "
        "Expected a row containing 'CODE' and year values (2020-2035)."
    )


def _detect_change_type_and_measure(
    rows: list[list[Any]],
    header_row: int,
    block_idx: int,
    block: dict[str, Any],
) -> tuple[str, str]:
    """Detect change type and measure for an intervention block.

    Searches the metadata rows above the grid and the header row itself.
    """
    raw_ct = ""
    raw_measure = ""

    # Strategy 1: Check header row for change type label
    # In real workbook, change type appears in col before CODE or at start
    first_year_col = min(block["year_cols"].values())
    for ci in range(max(0, first_year_col - 8), first_year_col):
        val = rows[header_row][ci] if ci < len(rows[header_row]) else None
        if isinstance(val, str) and val.strip():
            s = val.strip().lower()
            if any(kw in s for kw in [
                "absolute", "percentage", "sector_loss", "sector loss",
            ]):
                raw_ct = val.strip()
                break

    # Strategy 2: Scan rows above header for CHANGE TYPE / INTERVENTION MEASURE
    for scan_row in range(max(0, header_row - 25), header_row):
        row_vals = rows[scan_row] if scan_row < len(rows) else []
        for ci, val in enumerate(row_vals):
            if not isinstance(val, str):
                continue
            s = val.strip().upper()

            # Find change type value near the block's year columns
            if "CHANGE TYPE" in s:
                # Look in neighboring cells for the value
                for offset in [1, -1, -4, -5]:
                    nci = ci + offset
                    if 0 <= nci < len(row_vals) and row_vals[nci]:
                        candidate = str(row_vals[nci]).strip()
                        if candidate and candidate.upper() != "CHANGE TYPE":
                            # Check if this is for our block
                            if block_idx == 0 or nci < first_year_col:
                                raw_ct = raw_ct or candidate

            if "INTERVENTION MEASURE" in s:
                # Look in next row or neighboring cells
                for offset in [1, -1, -4, -5]:
                    nci = ci + offset
                    if 0 <= nci < len(row_vals) and row_vals[nci]:
                        candidate = str(row_vals[nci]).strip()
                        if candidate and "MEASURE" not in candidate.upper():
                            raw_measure = raw_measure or candidate
                # Also check the row below
                next_row = scan_row + 1
                if next_row < len(rows):
                    for nci in range(max(0, ci - 5), min(len(rows[next_row]), ci + 5)):
                        nval = rows[next_row][nci]
                        if nval and isinstance(nval, str):
                            candidate = nval.strip()
                            if candidate and "MEASURE" not in candidate.upper():
                                raw_measure = raw_measure or candidate

    return raw_ct, raw_measure


def _detect_intervention_labels(
    rows: list[list[Any]],
    header_row: int,
    blocks: list[dict[str, Any]],
) -> list[str]:
    """Detect intervention labels (INTERVENTION A, B, etc.).

    Finds the closest INTERVENTION label to each block's year columns.
    """
    # Collect all candidate labels from rows above header
    candidates: list[tuple[int, str]] = []  # (col_idx, label_text)
    for scan_row in range(max(0, header_row - 25), header_row):
        row_vals = rows[scan_row] if scan_row < len(rows) else []
        for ci, val in enumerate(row_vals):
            if isinstance(val, str) and "INTERVENTION" in val.upper():
                candidates.append((ci, val.strip()))

    labels = []
    for block_idx, block in enumerate(blocks):
        first_year_col = min(block["year_cols"].values())
        default_label = f"INTERVENTION {'ABCDEFGH'[block_idx]}"

        if not candidates:
            labels.append(default_label)
            continue

        # Find closest candidate label to this block's first year column
        best_dist = float("inf")
        best_label = default_label
        for ci, text in candidates:
            dist = abs(ci - first_year_col)
            if dist < best_dist:
                best_dist = dist
                best_label = text

        labels.append(best_label if best_dist < 15 else default_label)

    return labels


# =====================================================================
# Row readers
# =====================================================================


def _read_rows_xlsx(path: str) -> list[list[Any]]:
    """Read all rows from the first (or intervention) sheet using openpyxl."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

    # Try to find an INTERVENTIONS sheet
    target_sheet = None
    for name in wb.sheetnames:
        if "INTERVENTION" in name.upper():
            target_sheet = name
            break
    if target_sheet is None:
        target_sheet = wb.sheetnames[0]

    ws = wb[target_sheet]
    rows: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    wb.close()
    return rows


def _read_rows_xlsb(path: str) -> list[list[Any]]:
    """Read all rows from the intervention sheet using pyxlsb."""
    import pyxlsb

    wb = pyxlsb.open_workbook(path)

    # Find intervention sheet
    target_sheet = None
    for name in wb.sheets:
        if "INTERVENTION" in name.upper():
            target_sheet = name
            break
    if target_sheet is None:
        target_sheet = wb.sheets[0]

    rows: list[list[Any]] = []
    with wb.get_sheet(target_sheet) as sheet:
        for row in sheet.rows():
            rows.append([cell.v for cell in row])

    return rows


# =====================================================================
# Parser class
# =====================================================================


class SGTemplateParser:
    """Parse Strategic Gears production Excel templates (.xlsb or .xlsx)."""

    def __init__(self, concordance: ConcordanceService) -> None:
        self._concordance = concordance

    def parse(self, excel_path: str) -> SGScenarioImport:
        """Parse an SG template file. Auto-detects .xlsb vs .xlsx."""
        ext = Path(excel_path).suffix.lower()
        if ext == ".xlsb":
            rows = _read_rows_xlsb(excel_path)
        elif ext in (".xlsx", ".xls"):
            rows = _read_rows_xlsx(excel_path)
        else:
            raise ValueError(f"Unsupported file extension: {ext}")

        return self._parse_rows(rows, excel_path)

    def _parse_rows(
        self,
        rows: list[list[Any]],
        source_path: str,
    ) -> SGScenarioImport:
        """Parse raw row data into structured scenario."""
        # Detect grid structure
        header_row_idx, blocks = _find_header_row(rows)

        # Detect labels and metadata
        labels = _detect_intervention_labels(rows, header_row_idx, blocks)

        # Extract years from first block
        all_years = sorted(blocks[0]["year_cols"].keys()) if blocks else []

        # Parse each intervention block
        interventions = []
        all_codes: set[str] = set()

        for block_idx, block in enumerate(blocks):
            label = labels[block_idx] if block_idx < len(labels) else (
                f"INTERVENTION {'ABCDEFGH'[block_idx]}"
            )

            raw_ct, raw_measure = _detect_change_type_and_measure(
                rows, header_row_idx, block_idx, block,
            )
            norm_ct = normalize_change_type(raw_ct) if raw_ct else "unknown"
            norm_measure = _normalize_measure(raw_measure)

            code_col = block["code_col"]
            sector_col = block["sector_col"]
            year_cols = block["year_cols"]

            shock_rows: list[SGShockRow] = []
            active_rows: list[SGShockRow] = []

            # Read data rows after header
            for data_row_idx in range(header_row_idx + 1, len(rows)):
                row = rows[data_row_idx]
                if code_col >= len(row):
                    continue

                raw_code = row[code_col] if code_col < len(row) else None
                if raw_code is None or raw_code == "":
                    continue

                code = normalize_sector_code(raw_code)
                if not code or not code[0].isdigit():
                    continue

                name = ""
                if sector_col < len(row) and row[sector_col]:
                    name = str(row[sector_col]).strip()

                # Read annual values
                annual: dict[int, float] = {}
                for yr, col in year_cols.items():
                    val = row[col] if col < len(row) else 0.0
                    try:
                        annual[yr] = parse_parenthetical_negative(val) if val else 0.0
                    except (ValueError, TypeError):
                        annual[yr] = 0.0

                has_nonzero = any(abs(v) > 1e-12 for v in annual.values())

                shock_row = SGShockRow(
                    intervention_label=label,
                    raw_change_type=raw_ct,
                    normalized_change_type=norm_ct,
                    raw_measure_label=raw_measure,
                    normalized_measure=norm_measure,
                    sector_code=code,
                    sector_name=name,
                    is_selected=has_nonzero,
                    annual_values=annual,
                )
                shock_rows.append(shock_row)
                if has_nonzero:
                    active_rows.append(shock_row)
                all_codes.add(code)

            intervention = SGIntervention(
                label=label,
                change_type=norm_ct,
                measure=norm_measure,
                rows=shock_rows,
                active_rows=active_rows,
            )
            interventions.append(intervention)

        return SGScenarioImport(
            scenario_name=Path(source_path).stem,
            interventions=interventions,
            years=all_years,
            all_sector_codes=sorted(all_codes),
            metadata={
                "source_file": str(source_path),
                "header_row": header_row_idx,
                "block_count": len(blocks),
            },
        )

    def translation_report(
        self,
        scenario: SGScenarioImport,
    ) -> TranslationReport:
        """Classify interventions by engine compatibility.

        Supported NOW: absolute_value + final_demand
        NOT YET supported: percentage, sector_loss, compensation_of_employees
        """
        supported: list[SGIntervention] = []
        unsupported: list[SGIntervention] = []
        warn_msgs: list[str] = []

        for interv in scenario.interventions:
            if not interv.active_rows:
                continue

            is_supported = (
                interv.change_type == "absolute_value"
                and interv.measure == "final_demand"
            )

            if is_supported:
                supported.append(interv)
            else:
                unsupported.append(interv)
                reasons = []
                if interv.change_type != "absolute_value":
                    reasons.append(
                        f"change_type '{interv.change_type}' not supported"
                    )
                if interv.measure != "final_demand":
                    reasons.append(
                        f"measure '{interv.measure}' not supported"
                    )
                warn_msgs.append(
                    f"{interv.label}: {', '.join(reasons)} "
                    f"({len(interv.active_rows)} active sectors)"
                )

        return TranslationReport(
            supported=supported,
            unsupported=unsupported,
            warnings=warn_msgs,
        )

    def to_shock_vectors(
        self,
        scenario: SGScenarioImport,
        target_granularity: str = "section",
    ) -> dict[int, np.ndarray]:
        """Convert SUPPORTED interventions to engine shock vectors.

        ONLY handles absolute_value + final_demand.
        For unsupported types, use translation_report() first.

        Args:
            scenario: Parsed scenario.
            target_granularity: 'section' -> 20-sector vectors via concordance.
                               'division' -> 84-sector vectors.

        Returns:
            {year: shock_vector} for years with non-zero shocks.

        Raises:
            ValueError: If no supported interventions found.
        """
        report = self.translation_report(scenario)
        if not report.supported:
            raise ValueError(
                "No supported interventions found. "
                f"Unsupported: {[i.label for i in report.unsupported]}. "
                "Only absolute_value + final_demand is currently supported."
            )

        # Aggregate division-level shocks per year
        year_div_shocks: dict[int, dict[str, float]] = {}
        for interv in report.supported:
            for row in interv.active_rows:
                for yr, val in row.annual_values.items():
                    if abs(val) < 1e-12:
                        continue
                    year_div_shocks.setdefault(yr, {})
                    year_div_shocks[yr][row.sector_code] = (
                        year_div_shocks[yr].get(row.sector_code, 0.0) + val
                    )

        if target_granularity == "section":
            section_codes = self._concordance.section_codes
            result: dict[int, np.ndarray] = {}
            for yr, div_shocks in sorted(year_div_shocks.items()):
                # Aggregate divisions to sections
                section_shocks = self._concordance.aggregate_division_vector(
                    div_shocks, method="sum",
                )
                vec = np.zeros(len(section_codes))
                for i, sc in enumerate(section_codes):
                    vec[i] = section_shocks.get(sc, 0.0)
                result[yr] = vec
            return result

        elif target_granularity == "division":
            div_codes = self._concordance.all_division_codes
            result_div: dict[int, np.ndarray] = {}
            for yr, div_shocks in sorted(year_div_shocks.items()):
                vec = np.zeros(len(div_codes))
                for i, dc in enumerate(div_codes):
                    vec[i] = div_shocks.get(dc, 0.0)
                result_div[yr] = vec
            return result_div

        else:
            raise ValueError(
                f"Unknown target_granularity: '{target_granularity}'. "
                "Use 'section' or 'division'."
            )
