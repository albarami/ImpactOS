"""Excel export with integrity signature — MVP-6 Section 12.3.

Generate a fully linked Excel workbook with:
- Run metadata (run_id, model_version, scenario_version) in a visible footer sheet
- Input vectors
- Linked formulas reproducing core calculations
- Integrity signature (SHA-256 hash of key ranges) in a hidden sheet

If the workbook is modified outside ImpactOS, the signature fails.
Deterministic — no LLM calls.
"""

import hashlib
import io

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def _compute_range_hash(ws: Worksheet, min_row: int, max_row: int, min_col: int, max_col: int) -> str:
    """Compute SHA-256 hash of a cell range's values."""
    h = hashlib.sha256()
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            h.update(str(cell.value).encode())
    return h.hexdigest()


class ExcelExporter:
    """Generate governed Excel workbook with integrity tracking."""

    def export(self, pack_data: dict) -> bytes:
        """Generate Excel workbook bytes from Decision Pack data."""
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # 1. Sector Impacts sheet
        self._write_sector_impacts(wb, pack_data.get("sector_impacts", []))

        # 2. Input Vectors sheet
        self._write_input_vectors(wb, pack_data.get("input_vectors", {}))

        # 3. Assumptions sheet
        self._write_assumptions(wb, pack_data.get("assumptions", []))

        # 4. Run Metadata sheet (visible footer)
        self._write_metadata(wb, pack_data)

        # Save and reload so cell value representations are stable
        # (openpyxl normalises numeric types on round-trip)
        buf = io.BytesIO()
        wb.save(buf)
        wb = load_workbook(io.BytesIO(buf.getvalue()))

        # 5. Integrity signature (hidden sheet) — computed on reloaded wb
        self._write_integrity(wb)

        buf2 = io.BytesIO()
        wb.save(buf2)
        return buf2.getvalue()

    def _write_sector_impacts(self, wb: Workbook, impacts: list[dict]) -> None:
        ws = wb.create_sheet("Sector Impacts")
        headers = [
            "Sector Code", "Sector Name", "Direct Impact",
            "Indirect Impact", "Total Impact", "Multiplier",
            "Domestic Share", "Import Leakage",
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)

        for row_idx, si in enumerate(impacts, 2):
            ws.cell(row=row_idx, column=1, value=si["sector_code"])
            ws.cell(row=row_idx, column=2, value=si["sector_name"])
            ws.cell(row=row_idx, column=3, value=si["direct_impact"])
            ws.cell(row=row_idx, column=4, value=si["indirect_impact"])
            # Linked formula: total = direct + indirect
            ws.cell(row=row_idx, column=5, value=f"=C{row_idx}+D{row_idx}")
            ws.cell(row=row_idx, column=6, value=si["multiplier"])
            ws.cell(row=row_idx, column=7, value=si["domestic_share"])
            ws.cell(row=row_idx, column=8, value=si["import_leakage"])

    def _write_input_vectors(self, wb: Workbook, vectors: dict) -> None:
        ws = wb.create_sheet("Input Vectors")
        ws.cell(row=1, column=1, value="Sector Code")
        ws.cell(row=1, column=2, value="Input Value")

        for row_idx, (sector, value) in enumerate(sorted(vectors.items()), 2):
            ws.cell(row=row_idx, column=1, value=sector)
            ws.cell(row=row_idx, column=2, value=value)

    def _write_assumptions(self, wb: Workbook, assumptions: list[dict]) -> None:
        ws = wb.create_sheet("Assumptions")
        headers = ["Name", "Value", "Range Min", "Range Max"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)

        for row_idx, a in enumerate(assumptions, 2):
            ws.cell(row=row_idx, column=1, value=a.get("name", ""))
            ws.cell(row=row_idx, column=2, value=a.get("value"))
            ws.cell(row=row_idx, column=3, value=a.get("range_min"))
            ws.cell(row=row_idx, column=4, value=a.get("range_max"))

    def _write_metadata(self, wb: Workbook, pack_data: dict) -> None:
        ws = wb.create_sheet("Run Metadata")
        metadata = [
            ("Run ID", pack_data.get("run_id", "")),
            ("Scenario Name", pack_data.get("scenario_name", "")),
            ("Base Year", pack_data.get("base_year", "")),
            ("Currency", pack_data.get("currency", "")),
            ("Model Version", pack_data.get("model_version_id", "")),
            ("Scenario Version", pack_data.get("scenario_version", "")),
        ]
        for row_idx, (label, value) in enumerate(metadata, 1):
            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=2, value=str(value) if value is not None else "")

    def _write_integrity(self, wb: Workbook) -> None:
        """Compute integrity hash over key data ranges and store in hidden sheet."""
        h = hashlib.sha256()

        # Hash sector impacts data
        if "Sector Impacts" in wb.sheetnames:
            ws = wb["Sector Impacts"]
            max_row = ws.max_row or 1
            range_hash = _compute_range_hash(ws, 1, max_row, 1, 8)
            h.update(range_hash.encode())

        # Hash input vectors data
        if "Input Vectors" in wb.sheetnames:
            ws = wb["Input Vectors"]
            max_row = ws.max_row or 1
            range_hash = _compute_range_hash(ws, 1, max_row, 1, 2)
            h.update(range_hash.encode())

        # Hash metadata
        if "Run Metadata" in wb.sheetnames:
            ws = wb["Run Metadata"]
            max_row = ws.max_row or 1
            range_hash = _compute_range_hash(ws, 1, max_row, 1, 2)
            h.update(range_hash.encode())

        signature = f"sha256:{h.hexdigest()}"

        # Write to hidden sheet
        ws_integrity = wb.create_sheet("_Integrity")
        ws_integrity.cell(row=1, column=1, value="Integrity Signature")
        ws_integrity.cell(row=1, column=2, value=signature)
        ws_integrity.sheet_state = "hidden"


class IntegrityChecker:
    """Verify Excel workbook integrity signature."""

    def verify(self, workbook_bytes: bytes) -> bool:
        """Check if the workbook's integrity signature is still valid.

        Returns True if unmodified, False if tampered.
        """
        wb = load_workbook(io.BytesIO(workbook_bytes))

        if "_Integrity" not in wb.sheetnames:
            return False

        stored_signature = wb["_Integrity"].cell(row=1, column=2).value
        if not stored_signature or not stored_signature.startswith("sha256:"):
            return False

        # Recompute the hash
        h = hashlib.sha256()

        if "Sector Impacts" in wb.sheetnames:
            ws = wb["Sector Impacts"]
            max_row = ws.max_row or 1
            range_hash = _compute_range_hash(ws, 1, max_row, 1, 8)
            h.update(range_hash.encode())

        if "Input Vectors" in wb.sheetnames:
            ws = wb["Input Vectors"]
            max_row = ws.max_row or 1
            range_hash = _compute_range_hash(ws, 1, max_row, 1, 2)
            h.update(range_hash.encode())

        if "Run Metadata" in wb.sheetnames:
            ws = wb["Run Metadata"]
            max_row = ws.max_row or 1
            range_hash = _compute_range_hash(ws, 1, max_row, 1, 2)
            h.update(range_hash.encode())

        computed = f"sha256:{h.hexdigest()}"
        return computed == stored_signature
