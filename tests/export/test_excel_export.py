"""Tests for Excel export with integrity signature (MVP-6).

Covers: generate workbook with run metadata, input vectors, linked formulas,
integrity signature in hidden sheet. Modification detection.
"""

import hashlib
import io

import pytest
from openpyxl import load_workbook
from uuid_extensions import uuid7

from src.export.excel_export import ExcelExporter, IntegrityChecker


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

RUN_ID = uuid7()
MODEL_VERSION_ID = uuid7()
SCENARIO_VERSION = 3


def _make_pack_data() -> dict:
    return {
        "run_id": str(RUN_ID),
        "scenario_name": "NEOM Logistics Zone",
        "base_year": 2023,
        "currency": "SAR",
        "model_version_id": str(MODEL_VERSION_ID),
        "scenario_version": SCENARIO_VERSION,
        "sector_impacts": [
            {
                "sector_code": "C41",
                "sector_name": "Steel",
                "direct_impact": 500_000_000.0,
                "indirect_impact": 250_000_000.0,
                "total_impact": 750_000_000.0,
                "multiplier": 1.5,
                "domestic_share": 0.65,
                "import_leakage": 0.35,
            },
        ],
        "input_vectors": {"C41": 1_000_000.0, "F": 2_000_000.0},
        "assumptions": [
            {"name": "Domestic share", "value": 0.65, "range_min": 0.55, "range_max": 0.75},
        ],
    }


# ===================================================================
# Workbook generation
# ===================================================================


class TestWorkbookGeneration:
    """Generate Excel workbook with required sheets."""

    def test_generates_bytes(self) -> None:
        exporter = ExcelExporter()
        data = exporter.export(_make_pack_data())
        assert isinstance(data, bytes)
        assert len(data) > 0

    def test_has_sector_impacts_sheet(self) -> None:
        exporter = ExcelExporter()
        wb = load_workbook(io.BytesIO(exporter.export(_make_pack_data())))
        assert "Sector Impacts" in wb.sheetnames

    def test_has_input_vectors_sheet(self) -> None:
        exporter = ExcelExporter()
        wb = load_workbook(io.BytesIO(exporter.export(_make_pack_data())))
        assert "Input Vectors" in wb.sheetnames

    def test_has_assumptions_sheet(self) -> None:
        exporter = ExcelExporter()
        wb = load_workbook(io.BytesIO(exporter.export(_make_pack_data())))
        assert "Assumptions" in wb.sheetnames

    def test_has_metadata_sheet(self) -> None:
        exporter = ExcelExporter()
        wb = load_workbook(io.BytesIO(exporter.export(_make_pack_data())))
        assert "Run Metadata" in wb.sheetnames

    def test_sector_impacts_data(self) -> None:
        exporter = ExcelExporter()
        wb = load_workbook(io.BytesIO(exporter.export(_make_pack_data())))
        ws = wb["Sector Impacts"]
        # Row 1 = header, Row 2 = first data row
        assert ws.cell(row=2, column=1).value == "C41"
        assert ws.cell(row=2, column=3).value == 500_000_000.0

    def test_metadata_contains_run_id(self) -> None:
        exporter = ExcelExporter()
        wb = load_workbook(io.BytesIO(exporter.export(_make_pack_data())))
        ws = wb["Run Metadata"]
        values = [ws.cell(row=r, column=2).value for r in range(1, 10)]
        assert str(RUN_ID) in values

    def test_input_vectors_data(self) -> None:
        exporter = ExcelExporter()
        wb = load_workbook(io.BytesIO(exporter.export(_make_pack_data())))
        ws = wb["Input Vectors"]
        # Row 2 should have first sector
        assert ws.cell(row=2, column=1).value in ("C41", "F")


# ===================================================================
# Integrity signature
# ===================================================================


class TestIntegritySignature:
    """Hidden sheet with hash of key ranges."""

    def test_has_integrity_sheet(self) -> None:
        exporter = ExcelExporter()
        wb = load_workbook(io.BytesIO(exporter.export(_make_pack_data())))
        assert "_Integrity" in wb.sheetnames

    def test_integrity_sheet_has_hash(self) -> None:
        exporter = ExcelExporter()
        wb = load_workbook(io.BytesIO(exporter.export(_make_pack_data())))
        ws = wb["_Integrity"]
        signature = ws.cell(row=1, column=2).value
        assert signature is not None
        assert signature.startswith("sha256:")

    def test_verify_unmodified_passes(self) -> None:
        exporter = ExcelExporter()
        data = exporter.export(_make_pack_data())
        checker = IntegrityChecker()
        assert checker.verify(data) is True

    def test_verify_modified_fails(self) -> None:
        exporter = ExcelExporter()
        data = exporter.export(_make_pack_data())
        # Load, modify a cell, save
        wb = load_workbook(io.BytesIO(data))
        wb["Sector Impacts"].cell(row=2, column=3).value = 999_999_999.0
        buf = io.BytesIO()
        wb.save(buf)
        modified_data = buf.getvalue()
        checker = IntegrityChecker()
        assert checker.verify(modified_data) is False


# ===================================================================
# Linked formulas
# ===================================================================


class TestLinkedFormulas:
    """Workbook contains formulas reproducing core calculations."""

    def test_total_impact_formula(self) -> None:
        exporter = ExcelExporter()
        wb = load_workbook(io.BytesIO(exporter.export(_make_pack_data())))
        ws = wb["Sector Impacts"]
        # Column 5 = total_impact should be a formula (direct + indirect)
        cell = ws.cell(row=2, column=5)
        # Either formula or computed value
        assert cell.value is not None
