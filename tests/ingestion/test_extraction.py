"""Tests for extraction service (MVP-2 Section 8.2/8.3/8.5).

Covers: CSV extraction, Excel extraction, DocumentGraph construction,
EvidenceSnippet generation per extracted table.
"""

import csv
import io
from uuid import UUID

import pytest
from openpyxl import Workbook
from uuid_extensions import uuid7

from src.models.document import (
    DocumentGraph,
    ExtractionMetadata,
    ExtractedTable,
)
from src.models.governance import EvidenceSnippet
from src.ingestion.extraction import ExtractionService


# ---------------------------------------------------------------------------
# Helpers: generate test file content
# ---------------------------------------------------------------------------


def _make_csv_bytes(rows: list[list[str]]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8")


def _make_xlsx_bytes(rows: list[list[str | int | float]], sheet_name: str = "Sheet1") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


VALID_CHECKSUM = "sha256:" + "a" * 64

BOQ_CSV_ROWS = [
    ["Description", "Quantity", "Unit", "Unit Price", "Total"],
    ["Structural Steel", "5000", "tonnes", "3500", "17500000"],
    ["Concrete", "20000", "m3", "450", "9000000"],
    ["Rebar", "3000", "tonnes", "4200", "12600000"],
]

BOQ_XLSX_ROWS: list[list[str | int | float]] = [
    ["Description", "Quantity", "Unit", "Unit Price", "Total"],
    ["Structural Steel", 5000, "tonnes", 3500, 17500000],
    ["Concrete", 20000, "m3", 450, 9000000],
]


# ===================================================================
# CSV extraction
# ===================================================================


class TestCSVExtraction:
    """Extract tables from CSV content."""

    def test_csv_returns_document_graph(self) -> None:
        content = _make_csv_bytes(BOQ_CSV_ROWS)
        doc_id = uuid7()
        svc = ExtractionService()

        graph = svc.extract_csv(doc_id=doc_id, content=content, doc_checksum=VALID_CHECKSUM)

        assert isinstance(graph, DocumentGraph)
        assert graph.document_id == doc_id

    def test_csv_extracts_one_table(self) -> None:
        content = _make_csv_bytes(BOQ_CSV_ROWS)
        svc = ExtractionService()
        graph = svc.extract_csv(doc_id=uuid7(), content=content, doc_checksum=VALID_CHECKSUM)

        assert len(graph.pages) == 1
        assert len(graph.pages[0].tables) == 1

    def test_csv_table_has_correct_cells(self) -> None:
        content = _make_csv_bytes(BOQ_CSV_ROWS)
        svc = ExtractionService()
        graph = svc.extract_csv(doc_id=uuid7(), content=content, doc_checksum=VALID_CHECKSUM)

        table = graph.pages[0].tables[0]
        # 4 rows x 5 cols = 20 cells
        assert len(table.cells) == 20

    def test_csv_cell_text_matches_input(self) -> None:
        content = _make_csv_bytes(BOQ_CSV_ROWS)
        svc = ExtractionService()
        graph = svc.extract_csv(doc_id=uuid7(), content=content, doc_checksum=VALID_CHECKSUM)

        table = graph.pages[0].tables[0]
        # Find the cell at row=1, col=0 => "Structural Steel"
        cell = next(c for c in table.cells if c.row == 1 and c.col == 0)
        assert cell.text == "Structural Steel"

    def test_csv_cells_have_bounding_boxes(self) -> None:
        content = _make_csv_bytes(BOQ_CSV_ROWS)
        svc = ExtractionService()
        graph = svc.extract_csv(doc_id=uuid7(), content=content, doc_checksum=VALID_CHECKSUM)

        table = graph.pages[0].tables[0]
        for cell in table.cells:
            assert cell.bbox.x0 >= 0.0
            assert cell.bbox.x1 <= 1.0
            assert cell.bbox.y0 >= 0.0
            assert cell.bbox.y1 <= 1.0
            assert cell.bbox.x1 > cell.bbox.x0
            assert cell.bbox.y1 > cell.bbox.y0

    def test_csv_confidence_is_one(self) -> None:
        """CSV data is exact — confidence should be 1.0."""
        content = _make_csv_bytes(BOQ_CSV_ROWS)
        svc = ExtractionService()
        graph = svc.extract_csv(doc_id=uuid7(), content=content, doc_checksum=VALID_CHECKSUM)

        table = graph.pages[0].tables[0]
        for cell in table.cells:
            assert cell.confidence == 1.0

    def test_csv_extraction_metadata(self) -> None:
        content = _make_csv_bytes(BOQ_CSV_ROWS)
        svc = ExtractionService()
        graph = svc.extract_csv(doc_id=uuid7(), content=content, doc_checksum=VALID_CHECKSUM)

        assert graph.extraction_metadata.engine == "csv-parser"
        assert graph.extraction_metadata.completed_at is not None


# ===================================================================
# Excel extraction
# ===================================================================


class TestExcelExtraction:
    """Extract tables from Excel content."""

    def test_xlsx_returns_document_graph(self) -> None:
        content = _make_xlsx_bytes(BOQ_XLSX_ROWS)
        svc = ExtractionService()
        graph = svc.extract_excel(doc_id=uuid7(), content=content, doc_checksum=VALID_CHECKSUM)

        assert isinstance(graph, DocumentGraph)

    def test_xlsx_extracts_one_table_per_sheet(self) -> None:
        content = _make_xlsx_bytes(BOQ_XLSX_ROWS)
        svc = ExtractionService()
        graph = svc.extract_excel(doc_id=uuid7(), content=content, doc_checksum=VALID_CHECKSUM)

        # One sheet => one page
        assert len(graph.pages) == 1
        assert len(graph.pages[0].tables) == 1

    def test_xlsx_cell_text_matches(self) -> None:
        content = _make_xlsx_bytes(BOQ_XLSX_ROWS)
        svc = ExtractionService()
        graph = svc.extract_excel(doc_id=uuid7(), content=content, doc_checksum=VALID_CHECKSUM)

        table = graph.pages[0].tables[0]
        cell = next(c for c in table.cells if c.row == 1 and c.col == 0)
        assert cell.text == "Structural Steel"

    def test_xlsx_numeric_values_as_strings(self) -> None:
        content = _make_xlsx_bytes(BOQ_XLSX_ROWS)
        svc = ExtractionService()
        graph = svc.extract_excel(doc_id=uuid7(), content=content, doc_checksum=VALID_CHECKSUM)

        table = graph.pages[0].tables[0]
        cell = next(c for c in table.cells if c.row == 1 and c.col == 1)
        assert cell.text == "5000"

    def test_xlsx_confidence_is_one(self) -> None:
        """Excel data is exact — confidence should be 1.0."""
        content = _make_xlsx_bytes(BOQ_XLSX_ROWS)
        svc = ExtractionService()
        graph = svc.extract_excel(doc_id=uuid7(), content=content, doc_checksum=VALID_CHECKSUM)

        table = graph.pages[0].tables[0]
        for cell in table.cells:
            assert cell.confidence == 1.0

    def test_xlsx_extraction_metadata(self) -> None:
        content = _make_xlsx_bytes(BOQ_XLSX_ROWS)
        svc = ExtractionService()
        graph = svc.extract_excel(doc_id=uuid7(), content=content, doc_checksum=VALID_CHECKSUM)

        assert graph.extraction_metadata.engine == "openpyxl"
        assert graph.extraction_metadata.completed_at is not None

    def test_xlsx_multi_sheet(self) -> None:
        """Multiple sheets each produce a page."""
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "BoQ1"
        ws1.append(["Item", "Qty"])
        ws1.append(["Steel", 100])

        ws2 = wb.create_sheet("BoQ2")
        ws2.append(["Item", "Qty"])
        ws2.append(["Concrete", 200])

        buf = io.BytesIO()
        wb.save(buf)
        content = buf.getvalue()

        svc = ExtractionService()
        graph = svc.extract_excel(doc_id=uuid7(), content=content, doc_checksum=VALID_CHECKSUM)

        assert len(graph.pages) == 2
        assert len(graph.pages[0].tables) == 1
        assert len(graph.pages[1].tables) == 1


# ===================================================================
# EvidenceSnippet generation
# ===================================================================


class TestEvidenceSnippetGeneration:
    """Generate EvidenceSnippets from extracted tables."""

    def test_generates_snippets_from_csv(self) -> None:
        content = _make_csv_bytes(BOQ_CSV_ROWS)
        svc = ExtractionService()
        graph = svc.extract_csv(doc_id=uuid7(), content=content, doc_checksum=VALID_CHECKSUM)
        doc_id = graph.document_id

        snippets = svc.generate_evidence_snippets(
            document_graph=graph,
            source_id=doc_id,
            doc_checksum=VALID_CHECKSUM,
        )
        assert len(snippets) > 0
        assert all(isinstance(s, EvidenceSnippet) for s in snippets)

    def test_snippet_has_page_and_bbox(self) -> None:
        content = _make_csv_bytes(BOQ_CSV_ROWS)
        svc = ExtractionService()
        graph = svc.extract_csv(doc_id=uuid7(), content=content, doc_checksum=VALID_CHECKSUM)

        snippets = svc.generate_evidence_snippets(
            document_graph=graph,
            source_id=graph.document_id,
            doc_checksum=VALID_CHECKSUM,
        )
        for snippet in snippets:
            assert snippet.page >= 0
            assert snippet.bbox.x0 >= 0.0
            assert snippet.bbox.x1 <= 1.0

    def test_snippet_stores_checksum(self) -> None:
        content = _make_csv_bytes(BOQ_CSV_ROWS)
        svc = ExtractionService()
        graph = svc.extract_csv(doc_id=uuid7(), content=content, doc_checksum=VALID_CHECKSUM)

        snippets = svc.generate_evidence_snippets(
            document_graph=graph,
            source_id=graph.document_id,
            doc_checksum=VALID_CHECKSUM,
        )
        for snippet in snippets:
            assert snippet.checksum == VALID_CHECKSUM

    def test_snippet_has_table_cell_ref(self) -> None:
        content = _make_csv_bytes(BOQ_CSV_ROWS)
        svc = ExtractionService()
        graph = svc.extract_csv(doc_id=uuid7(), content=content, doc_checksum=VALID_CHECKSUM)

        snippets = svc.generate_evidence_snippets(
            document_graph=graph,
            source_id=graph.document_id,
            doc_checksum=VALID_CHECKSUM,
        )
        for snippet in snippets:
            assert snippet.table_cell_ref is not None

    def test_snippet_per_data_row(self) -> None:
        """One snippet per data row (skipping header)."""
        content = _make_csv_bytes(BOQ_CSV_ROWS)
        svc = ExtractionService()
        graph = svc.extract_csv(doc_id=uuid7(), content=content, doc_checksum=VALID_CHECKSUM)

        snippets = svc.generate_evidence_snippets(
            document_graph=graph,
            source_id=graph.document_id,
            doc_checksum=VALID_CHECKSUM,
        )
        # 3 data rows (excluding header)
        assert len(snippets) == 3
