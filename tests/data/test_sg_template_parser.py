"""Tests for SG Template Parser (D-2).

Validates:
    src/data/sg_template_parser.py — Excel template parsing, normalization,
    shock vector conversion.

Test fixtures are programmatically-generated .xlsx files (NOT .xlsb).
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import numpy as np
import openpyxl
import pytest

from src.data.concordance import ConcordanceService
from src.data.sg_template_parser import (
    SGScenarioImport,
    SGTemplateParser,
    TranslationReport,
    normalize_change_type,
    normalize_sector_code,
    parse_parenthetical_negative,
)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
DATA_DIR = Path(__file__).resolve().parent.parent.parent / "data" / "curated"
CONCORDANCE_PATH = DATA_DIR / "concordance_section_division.json"
WEIGHTS_PATH = DATA_DIR / "division_output_weights_sg_2018.json"


# ---------------------------------------------------------------------------
# Test fixture: programmatically generated .xlsx template
# ---------------------------------------------------------------------------

def _build_sg_template_xlsx(
    path: Path,
    *,
    interventions: int = 2,
    change_types: list[str] | None = None,
    measures: list[str] | None = None,
    sector_data: list[tuple[Any, str, list[float]]] | None = None,
    sector_data_b: list[list[float]] | None = None,
) -> Path:
    """Generate a minimal SG-style Excel template matching real SG layout.

    Real layout (from actual workbook analysis):
        Row 2 (1-indexed): "INTERVENTION A" at col G, "INTERVENTION B" at col P
        Row 4: change type value at col B, "CHANGE TYPE" at col F
        Row 6: "INTERVENTION MEASURE" at col F
        Row 7: measure value at col B
        Row 21: Repeat intervention labels
        Row 23: Grid header: change_type at A, CODE at E, SECTOR at F,
                years 2025-2030 at G-L; change_type_B at N, years at P-U
        Row 24+: Sector data rows

    For simplicity, we use a streamlined version that the dynamic parser
    can still detect via CODE/SECTOR/year pattern matching.

    Args:
        sector_data_b: Optional per-sector values for Intervention B.
            Must be same length as sector_data. If None, zeros are used.
    """
    if change_types is None:
        change_types = ["Absolute_Value", "Percentage"]
    if measures is None:
        measures = ["", "Compensation of employees"]
    if sector_data is None:
        # Default: 3 sectors
        sector_data = [
            (6.0, "Extraction of crude petroleum", [500.0, 500.0, 0, 0, 0, 0]),
            ("41", "Construction of buildings", [0, 200.0, 300.0, 300.0, 0, 0]),
            (" 10 ", "Manufacture of food products", [0, 0, 0, 0, 0, 0]),
        ]

    years = [2025, 2026, 2027, 2028, 2029, 2030]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "INTERVENTIONS - SCENARIO 1"

    # --- Intervention A ---
    # Row 2: Intervention label (openpyxl is 1-indexed)
    ws.cell(row=2, column=7, value="INTERVENTION A")

    # Row 4: Change type
    ct_a = change_types[0] if change_types else "Absolute_Value"
    ws.cell(row=4, column=2, value=ct_a)
    ws.cell(row=4, column=6, value="CHANGE TYPE")

    # Row 6: Intervention measure label
    ws.cell(row=6, column=6, value="INTERVENTION MEASURE")

    # Row 7: Measure value
    measure_a = measures[0] if measures else ""
    ws.cell(row=7, column=2, value=measure_a)

    # Row 21: Repeat labels
    ws.cell(row=21, column=7, value="INTERVENTION A")

    # Row 23: Grid header for A
    ws.cell(row=23, column=1, value=ct_a.replace("_", " ").title())  # e.g. "Absolute Value"
    ws.cell(row=23, column=5, value="CODE")
    ws.cell(row=23, column=6, value="SECTOR")
    for yi, yr in enumerate(years):
        ws.cell(row=23, column=7 + yi, value=yr)

    # Row 24+: Sector data for A
    for ri, (code, name, values) in enumerate(sector_data):
        ws.cell(row=24 + ri, column=5, value=code)
        ws.cell(row=24 + ri, column=6, value=name)
        for vi, val in enumerate(values):
            ws.cell(row=24 + ri, column=7 + vi, value=val)

    # --- Intervention B (if requested) ---
    if interventions >= 2 and len(change_types) >= 2:
        ct_b = change_types[1]
        measure_b = measures[1] if len(measures) > 1 else ""

        ws.cell(row=2, column=16, value="INTERVENTION B")
        ws.cell(row=4, column=16, value=ct_b)
        ws.cell(row=7, column=16, value=measure_b)
        ws.cell(row=21, column=16, value="INTERVENTION B")

        # Grid header for B
        ws.cell(row=23, column=14, value=ct_b.replace("_", " ").title())
        for yi, yr in enumerate(years):
            ws.cell(row=23, column=16 + yi, value=yr)

        # Sector data for B (same sectors, zero values unless sector_data_b given)
        for ri, (code, _name, _values) in enumerate(sector_data):
            ws.cell(row=24 + ri, column=14, value=f" - {normalize_sector_code(code)}")
            ws.cell(row=24 + ri, column=15, value="")  # selection marker
            b_vals = sector_data_b[ri] if sector_data_b else [0.0] * len(years)
            for vi in range(len(years)):
                ws.cell(row=24 + ri, column=16 + vi, value=b_vals[vi])

    wb.save(str(path))
    return path


# ===================================================================
# TestCodeNormalization
# ===================================================================


class TestCodeNormalization:
    """Sector code and change type normalization."""

    def test_float_to_string(self) -> None:
        """10.0 -> '10'."""
        assert normalize_sector_code(10.0) == "10"

    def test_string_preserved(self) -> None:
        """'06' -> '06'."""
        assert normalize_sector_code("06") == "06"

    def test_whitespace_trimmed(self) -> None:
        """' 06 ' -> '06'."""
        assert normalize_sector_code(" 06 ") == "06"

    def test_int_zero_padded(self) -> None:
        """6 -> '06'."""
        assert normalize_sector_code(6) == "06"

    def test_float_zero_padded(self) -> None:
        """6.0 -> '06'."""
        assert normalize_sector_code(6.0) == "06"

    def test_two_digit_int(self) -> None:
        """41 -> '41' (no extra padding)."""
        assert normalize_sector_code(41) == "41"

    def test_change_type_absolute_value_underscore(self) -> None:
        """'Absolute_Value' -> 'absolute_value'."""
        assert normalize_change_type("Absolute_Value") == "absolute_value"

    def test_change_type_absolute_value_space(self) -> None:
        """'Absolute Value' -> 'absolute_value'."""
        assert normalize_change_type("Absolute Value") == "absolute_value"

    def test_change_type_percentage(self) -> None:
        """'Percentage' -> 'percentage'."""
        assert normalize_change_type("Percentage") == "percentage"

    def test_change_type_sector_loss(self) -> None:
        """'Sector Loss' -> 'sector_loss'."""
        assert normalize_change_type("Sector Loss") == "sector_loss"

    def test_change_type_whitespace(self) -> None:
        """'  Percentage ' -> 'percentage'."""
        assert normalize_change_type("  Percentage ") == "percentage"

    def test_parenthetical_negative(self) -> None:
        """(0.05) -> -0.05."""
        assert parse_parenthetical_negative("(0.05)") == pytest.approx(-0.05)

    def test_parenthetical_positive(self) -> None:
        """0.05 -> 0.05."""
        assert parse_parenthetical_negative(0.05) == pytest.approx(0.05)

    def test_parenthetical_string_negative(self) -> None:
        """'-0.05' -> -0.05."""
        assert parse_parenthetical_negative("-0.05") == pytest.approx(-0.05)

    def test_parenthetical_zero(self) -> None:
        """0 -> 0.0."""
        assert parse_parenthetical_negative(0) == pytest.approx(0.0)


# ===================================================================
# TestSGTemplateParser
# ===================================================================


@pytest.mark.skipif(
    not CONCORDANCE_PATH.exists(),
    reason="Concordance JSON not generated",
)
class TestSGTemplateParser:
    """SG template parsing tests."""

    def _make_parser(self) -> SGTemplateParser:
        svc = ConcordanceService(str(CONCORDANCE_PATH), str(WEIGHTS_PATH))
        return SGTemplateParser(svc)

    def test_parse_basic_template(self, tmp_path: Path) -> None:
        """Basic .xlsx fixture parses to SGScenarioImport."""
        xlsx = _build_sg_template_xlsx(tmp_path / "test.xlsx")
        parser = self._make_parser()
        result = parser.parse(str(xlsx))
        assert isinstance(result, SGScenarioImport)
        assert len(result.interventions) >= 1
        assert result.years == [2025, 2026, 2027, 2028, 2029, 2030]

    def test_detects_absolute_value_shocks(self, tmp_path: Path) -> None:
        """Correctly identifies absolute_value change type."""
        xlsx = _build_sg_template_xlsx(
            tmp_path / "test.xlsx",
            interventions=1,
            change_types=["Absolute_Value"],
            measures=[""],
        )
        parser = self._make_parser()
        result = parser.parse(str(xlsx))
        assert result.interventions[0].change_type == "absolute_value"
        assert result.interventions[0].measure == "final_demand"

    def test_detects_percentage_shocks(self, tmp_path: Path) -> None:
        """Percentage type parsed correctly."""
        xlsx = _build_sg_template_xlsx(
            tmp_path / "test.xlsx",
            interventions=1,
            change_types=["Percentage"],
            measures=[""],
        )
        parser = self._make_parser()
        result = parser.parse(str(xlsx))
        assert result.interventions[0].change_type == "percentage"

    def test_multi_intervention_parsed(self, tmp_path: Path) -> None:
        """Both INTERVENTION A and B extracted."""
        xlsx = _build_sg_template_xlsx(
            tmp_path / "test.xlsx",
            interventions=2,
        )
        parser = self._make_parser()
        result = parser.parse(str(xlsx))
        assert len(result.interventions) == 2
        labels = [i.label for i in result.interventions]
        assert "INTERVENTION A" in labels
        assert "INTERVENTION B" in labels

    def test_multi_year_phasing_preserved(self, tmp_path: Path) -> None:
        """Year-by-year values are correct."""
        sector_data = [
            ("06", "Oil", [100.0, 200.0, 300.0, 400.0, 500.0, 600.0]),
        ]
        xlsx = _build_sg_template_xlsx(
            tmp_path / "test.xlsx",
            interventions=1,
            change_types=["Absolute_Value"],
            measures=[""],
            sector_data=sector_data,
        )
        parser = self._make_parser()
        result = parser.parse(str(xlsx))
        active = result.interventions[0].active_rows
        assert len(active) == 1
        row = active[0]
        assert row.annual_values[2025] == pytest.approx(100.0)
        assert row.annual_values[2030] == pytest.approx(600.0)

    def test_empty_intervention_skipped(self, tmp_path: Path) -> None:
        """All-zero sector -> not in active_rows."""
        sector_data = [
            ("06", "Oil", [0, 0, 0, 0, 0, 0]),
        ]
        xlsx = _build_sg_template_xlsx(
            tmp_path / "test.xlsx",
            interventions=1,
            change_types=["Absolute_Value"],
            measures=[""],
            sector_data=sector_data,
        )
        parser = self._make_parser()
        result = parser.parse(str(xlsx))
        assert len(result.interventions[0].active_rows) == 0

    def test_sector_code_normalization_in_parse(self, tmp_path: Path) -> None:
        """Codes like 6.0 and ' 10 ' are normalized in parsed output."""
        sector_data = [
            (6.0, "Oil", [500.0, 0, 0, 0, 0, 0]),
            (" 10 ", "Food", [100.0, 0, 0, 0, 0, 0]),
        ]
        xlsx = _build_sg_template_xlsx(
            tmp_path / "test.xlsx",
            interventions=1,
            change_types=["Absolute_Value"],
            measures=[""],
            sector_data=sector_data,
        )
        parser = self._make_parser()
        result = parser.parse(str(xlsx))
        codes = {r.sector_code for r in result.interventions[0].active_rows}
        assert "06" in codes
        assert "10" in codes

    def test_xlsb_backend_selected(self) -> None:
        """A .xlsb path triggers the xlsb code path (ValueError if file missing)."""
        parser = self._make_parser()
        with pytest.raises((FileNotFoundError, ValueError, OSError)):
            parser.parse("nonexistent.xlsb")

    def test_unsupported_extension_raises(self) -> None:
        """Unsupported file extension raises ValueError."""
        parser = self._make_parser()
        with pytest.raises(ValueError, match="[Uu]nsupported"):
            parser.parse("file.csv")


# ===================================================================
# TestTranslationReport
# ===================================================================


@pytest.mark.skipif(
    not CONCORDANCE_PATH.exists(),
    reason="Concordance JSON not generated",
)
class TestTranslationReport:
    """Translation report classifying engine compatibility."""

    def _make_parser(self) -> SGTemplateParser:
        svc = ConcordanceService(str(CONCORDANCE_PATH), str(WEIGHTS_PATH))
        return SGTemplateParser(svc)

    def test_absolute_final_demand_supported(self, tmp_path: Path) -> None:
        """absolute_value + final_demand -> supported."""
        xlsx = _build_sg_template_xlsx(
            tmp_path / "test.xlsx",
            interventions=1,
            change_types=["Absolute_Value"],
            measures=[""],
        )
        parser = self._make_parser()
        scenario = parser.parse(str(xlsx))
        report = parser.translation_report(scenario)
        assert isinstance(report, TranslationReport)
        assert len(report.supported) >= 1
        assert report.supported[0].change_type == "absolute_value"

    def test_percentage_unsupported(self, tmp_path: Path) -> None:
        """percentage -> unsupported list."""
        xlsx = _build_sg_template_xlsx(
            tmp_path / "test.xlsx",
            interventions=1,
            change_types=["Percentage"],
            measures=[""],
        )
        parser = self._make_parser()
        scenario = parser.parse(str(xlsx))
        report = parser.translation_report(scenario)
        assert len(report.unsupported) >= 1

    def test_compensation_unsupported(self, tmp_path: Path) -> None:
        """compensation_of_employees -> unsupported."""
        xlsx = _build_sg_template_xlsx(
            tmp_path / "test.xlsx",
            interventions=1,
            change_types=["Absolute_Value"],
            measures=["Compensation of employees"],
        )
        parser = self._make_parser()
        scenario = parser.parse(str(xlsx))
        report = parser.translation_report(scenario)
        assert len(report.unsupported) >= 1

    def test_warnings_generated(self, tmp_path: Path) -> None:
        """Clear warning messages for unsupported interventions."""
        sector_data = [
            ("06", "Oil", [500.0, 0, 0, 0, 0, 0]),
            ("41", "Construction", [0, 300.0, 0, 0, 0, 0]),
        ]
        # Intervention B needs non-zero values so active_rows triggers warnings
        sector_data_b = [
            [100.0, 0, 0, 0, 0, 0],
            [0, 50.0, 0, 0, 0, 0],
        ]
        xlsx = _build_sg_template_xlsx(
            tmp_path / "test.xlsx",
            interventions=2,
            change_types=["Absolute_Value", "Percentage"],
            measures=["", ""],
            sector_data=sector_data,
            sector_data_b=sector_data_b,
        )
        parser = self._make_parser()
        scenario = parser.parse(str(xlsx))
        report = parser.translation_report(scenario)
        assert len(report.warnings) >= 1
        assert any("percentage" in w.lower() for w in report.warnings)


# ===================================================================
# TestShockVectorConversion
# ===================================================================


@pytest.mark.skipif(
    not CONCORDANCE_PATH.exists(),
    reason="Concordance JSON not generated",
)
class TestShockVectorConversion:
    """Shock vector generation from parsed template."""

    def _make_parser(self) -> SGTemplateParser:
        svc = ConcordanceService(str(CONCORDANCE_PATH), str(WEIGHTS_PATH))
        return SGTemplateParser(svc)

    def test_to_section_vectors(self, tmp_path: Path) -> None:
        """84-division -> 20-section aggregation."""
        sector_data = [
            ("06", "Oil", [500.0, 0, 0, 0, 0, 0]),
            ("41", "Construction", [0, 300.0, 0, 0, 0, 0]),
        ]
        xlsx = _build_sg_template_xlsx(
            tmp_path / "test.xlsx",
            interventions=1,
            change_types=["Absolute_Value"],
            measures=[""],
            sector_data=sector_data,
        )
        parser = self._make_parser()
        scenario = parser.parse(str(xlsx))
        vectors = parser.to_shock_vectors(scenario, target_granularity="section")
        assert 2025 in vectors
        assert 2026 in vectors
        # 2025: only oil (section B) should be non-zero
        v2025 = vectors[2025]
        assert isinstance(v2025, np.ndarray)
        assert len(v2025) == 20

    def test_per_year_vectors(self, tmp_path: Path) -> None:
        """Produces {2025: [...], 2026: [...], ...} for all active years."""
        sector_data = [
            ("06", "Oil", [100.0, 200.0, 300.0, 0, 0, 0]),
        ]
        xlsx = _build_sg_template_xlsx(
            tmp_path / "test.xlsx",
            interventions=1,
            change_types=["Absolute_Value"],
            measures=[""],
            sector_data=sector_data,
        )
        parser = self._make_parser()
        scenario = parser.parse(str(xlsx))
        vectors = parser.to_shock_vectors(scenario, target_granularity="section")
        # Should have vectors for years with non-zero data
        assert 2025 in vectors
        assert 2026 in vectors
        assert 2027 in vectors

    def test_unsupported_raises(self, tmp_path: Path) -> None:
        """Passing scenario with only unsupported interventions raises."""
        xlsx = _build_sg_template_xlsx(
            tmp_path / "test.xlsx",
            interventions=1,
            change_types=["Percentage"],
            measures=[""],
            sector_data=[("06", "Oil", [500.0, 0, 0, 0, 0, 0])],
        )
        parser = self._make_parser()
        scenario = parser.parse(str(xlsx))
        with pytest.raises(ValueError, match="[Nn]o supported"):
            parser.to_shock_vectors(scenario, target_granularity="section")
