"""Generate SG test fixtures for Sprint 18 adapter and parity tests.

Creates:
  - sg_3sector_model.xlsx   — 3-sector I-O workbook mimicking SG production layout
  - sg_parity_benchmark_v1.json — golden parity benchmark computed by LeontiefSolver

Run: python tests/fixtures/generate_sg_fixture.py
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import numpy as np
from openpyxl import Workbook

# ---------------------------------------------------------------------------
# Ensure project root is importable
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(PROJECT_ROOT))

from src.engine.leontief import LeontiefSolver  # noqa: E402
from src.engine.model_store import ModelStore  # noqa: E402

# ---------------------------------------------------------------------------
# Fixed model data
# ---------------------------------------------------------------------------
Z: list[list[float]] = [
    [150.0, 500.0, 100.0],
    [200.0, 100.0, 300.0],
    [50.0, 200.0, 50.0],
]
X: list[float] = [1000.0, 2000.0, 1500.0]
SECTOR_CODES: list[str] = ["01", "02", "03"]
SECTOR_NAMES: dict[str, str] = {
    "01": "Agriculture",
    "02": "Manufacturing",
    "03": "Services",
}
BASE_YEAR: int = 2024

FINAL_DEMAND_F: list[list[float]] = [
    [100.0, 50.0],
    [200.0, 150.0],
    [300.0, 100.0],
]
IMPORTS_VECTOR: list[float] = [120.0, 350.0, 80.0]
COMPENSATION: list[float] = [200.0, 400.0, 500.0]
GOS: list[float] = [150.0, 300.0, 350.0]
TAXES: list[float] = [50.0, 100.0, 70.0]

# Scenario parameters for parity benchmark
SHOCK_VECTOR: list[float] = [50.0, 0.0, 25.0]
JOBS_COEFF: list[float] = [10.0, 5.0, 8.0]

FIXTURES_DIR = Path(__file__).resolve().parent


def generate_xlsx(path: Path) -> None:
    """Write the 3-sector SG-format .xlsx workbook."""
    wb = Workbook()

    # ---- Sheet: IO_MODEL ----
    ws_io = wb.active
    ws_io.title = "IO_MODEL"
    ws_io.append(["SG IO Model - Test Fixture"])
    ws_io.append(["BASE_YEAR", BASE_YEAR])
    ws_io.append([])  # row 3 blank
    ws_io.append(["CODE", "SECTOR"] + SECTOR_CODES)  # row 4 header
    for i, code in enumerate(SECTOR_CODES):
        ws_io.append([code, SECTOR_NAMES[code]] + Z[i])  # rows 5-7
    ws_io.append([])  # row 8 blank
    ws_io.append(["TOTAL_OUTPUT", "Total Output"] + X)  # row 9

    # ---- Sheet: FINAL_DEMAND ----
    ws_fd = wb.create_sheet("FINAL_DEMAND")
    ws_fd.append(["CODE", "SECTOR", "Household", "Government"])
    for i, code in enumerate(SECTOR_CODES):
        ws_fd.append([code, SECTOR_NAMES[code]] + FINAL_DEMAND_F[i])

    # ---- Sheet: IMPORTS ----
    ws_imp = wb.create_sheet("IMPORTS")
    ws_imp.append(["CODE", "SECTOR", "Imports"])
    for i, code in enumerate(SECTOR_CODES):
        ws_imp.append([code, SECTOR_NAMES[code], IMPORTS_VECTOR[i]])

    # ---- Sheet: VALUE_ADDED ----
    ws_va = wb.create_sheet("VALUE_ADDED")
    ws_va.append(["CODE", "SECTOR", "Compensation", "GOS", "Taxes"])
    for i, code in enumerate(SECTOR_CODES):
        ws_va.append([
            code,
            SECTOR_NAMES[code],
            COMPENSATION[i],
            GOS[i],
            TAXES[i],
        ])

    wb.save(path)
    print(f"  wrote {path}")


def generate_benchmark_json(path: Path) -> None:
    """Compute parity benchmark via LeontiefSolver and write JSON."""
    store = ModelStore()
    mv = store.register(
        Z=np.array(Z, dtype=np.float64),
        x=np.array(X, dtype=np.float64),
        sector_codes=SECTOR_CODES,
        base_year=BASE_YEAR,
        source="sg_fixture_generator",
    )
    loaded = store.get(mv.model_version_id)

    solver = LeontiefSolver()
    shock = np.array(SHOCK_VECTOR, dtype=np.float64)
    result = solver.solve(loaded_model=loaded, delta_d=shock)

    # Scalar expected outputs — parity gate compares sums, not per-sector vectors
    total_output = float(np.sum(result.delta_x_total))

    # Employment: sum(delta_x_total * jobs_coeff) — matches parity_gate._compute_actual_outputs
    jobs = np.array(JOBS_COEFF, dtype=np.float64)
    employment = float(np.sum(result.delta_x_total * jobs))

    benchmark = {
        "benchmark_id": "sg_3sector_golden_v1",
        "description": "3-sector toy model parity check",
        "model": {
            "Z": Z,
            "x": X,
            "sector_codes": SECTOR_CODES,
            "base_year": BASE_YEAR,
        },
        "scenario": {
            "shock_vector": SHOCK_VECTOR,
            "jobs_coeff": JOBS_COEFF,
        },
        "expected_outputs": {
            "total_output": round(total_output, 10),
            "employment": round(employment, 10),
        },
        "tolerance": 0.001,
    }

    path.write_text(json.dumps(benchmark, indent=2) + "\n", encoding="utf-8")
    print(f"  wrote {path}")


def main() -> None:
    """Generate all fixtures."""
    print("Generating SG test fixtures...")

    xlsx_path = FIXTURES_DIR / "sg_3sector_model.xlsx"
    json_path = FIXTURES_DIR / "sg_parity_benchmark_v1.json"

    generate_xlsx(xlsx_path)
    generate_benchmark_json(json_path)

    print("Done.")


if __name__ == "__main__":
    main()
